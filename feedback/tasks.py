"""
Celery task: generate a multi-sheet Excel feedback report and email it.

Sheets produced:
  1. Summary        – today's counts + all-time totals per category
  2. Club           – ALL Club feedback rows (till date)
  3. Spa            – ALL Spa feedback rows (till date)
  4. Hotel          – ALL Hotel feedback rows (till date)
  5. Restaurant     – ALL Restaurant feedback rows (till date)

Usage:
  # Trigger manually from Django shell
  from feedback.tasks import send_feedback_report
  send_feedback_report.delay()

  # Runs automatically daily at 8 AM IST via Celery Beat (see sdcorp/celery.py)
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timedelta

from celery import shared_task
from django.conf import settings
from django.core.mail import EmailMessage
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import FEEDBACK_MODELS

logger = logging.getLogger(__name__)

FEEDBACK_LABELS = {
    "club": "Club Feedback",
    "spa": "Spa Feedback",
    "hotel": "Hotel Feedback",
    "restaurant": "Resident Dining Feedback",
}

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
SUMMARY_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")


def _display_value(obj, field):
    """Return a human-readable cell value for a model field."""
    value = getattr(obj, field.name)
    if field.choices:
        return getattr(obj, f"get_{field.name}_display")()
    if isinstance(value, datetime):
        return timezone.localtime(value).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, list):
        return ", ".join(str(v) for v in value) if value else ""
    if value is None:
        return ""
    return value


def _style_header_row(ws, col_count):
    """Apply styling to the first row (headers)."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _write_data_sheet(wb, sheet_name, model):
    """Write ALL rows for one feedback type into a dedicated sheet."""
    ws = wb.create_sheet(title=sheet_name)
    fields = list(model._meta.fields)

    headers = [f.verbose_name.title() for f in fields]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    row_count = 0
    for obj in model.objects.all().order_by("-submitted_at").iterator():
        ws.append([_display_value(obj, f) for f in fields])
        row_count += 1

    _auto_width(ws)
    return row_count


def _write_summary_sheet(wb, total_counts, report_date):
    """First sheet: today's counts + all-time totals."""
    ws = wb.active
    ws.title = "Summary"

    since = report_date - timedelta(days=1)

    today_counts = {}
    for slug, model in FEEDBACK_MODELS.items():
        today_counts[slug] = model.objects.filter(submitted_at__gte=since).count()

    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Imperial Club — Feedback Report"
    title_cell.font = SUMMARY_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws["A3"] = "Report Date"
    ws["B3"] = report_date.strftime("%d %b %Y %I:%M %p")
    ws["A3"].font = Font(bold=True)

    summary_headers = ["Category", "Today's Responses", "Total Responses (Till Date)"]
    ws.append([])  # row 4 spacer
    ws.append(summary_headers)
    header_row = ws.max_row
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    # clear auto-styled row 1 header fill (it's the title row)
    for col_idx in range(1, len(summary_headers) + 1):
        ws.cell(row=1, column=col_idx).fill = PatternFill()
        ws.cell(row=1, column=col_idx).border = Border()

    today_total = 0
    all_total = 0
    for slug in FEEDBACK_MODELS:
        label = FEEDBACK_LABELS.get(slug, slug.title())
        t_count = today_counts.get(slug, 0)
        a_count = total_counts.get(slug, 0)
        ws.append([label, t_count, a_count])
        today_total += t_count
        all_total += a_count

    ws.append([])
    ws.append(["Total", today_total, all_total])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)

    _auto_width(ws)


def _build_workbook(report_date):
    """Build the full Excel workbook and return bytes."""
    wb = Workbook()
    counts = {}

    for slug, model in FEEDBACK_MODELS.items():
        sheet_name = FEEDBACK_LABELS.get(slug, slug.title())
        row_count = _write_data_sheet(wb, sheet_name, model)
        counts[slug] = row_count

    _write_summary_sheet(wb, counts, report_date)

    # Move Summary to the first position
    summary_ws = wb["Summary"]
    wb.move_sheet(summary_ws, offset=-len(wb.sheetnames) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@shared_task(bind=True, max_retries=3, default_retry_delay=300)
def send_feedback_report(self, **kwargs):
    """
    Build a multi-sheet Excel report with ALL feedback till date and email it.
    Runs daily at 8 AM IST via Celery Beat.
    """
    now = timezone.localtime(timezone.now())

    try:
        excel_bytes = _build_workbook(now)
    except Exception as exc:
        logger.exception("Failed to build feedback workbook")
        raise self.retry(exc=exc)

    date_label = now.strftime("%d-%b-%Y")
    filename = f"Imperial_Club_Feedback_{date_label}.xlsx"

    subject = f"Imperial Club — Daily Feedback Report ({date_label})"
    body = (
        f"Dear Team,\n\n"
        f"Please find attached the complete feedback report as of {date_label}.\n\n"
        f"This report contains all responses received till date "
        f"across all feedback categories (Club, Spa, Hotel, Restaurant).\n\n"
        f"The Excel file includes:\n"
        f"  • Summary sheet with total response counts\n"
        f"  • Individual sheets for each feedback category\n\n"
        f"Regards,\n"
        f"Imperial Club Feedback System"
    )

    recipients = getattr(settings, "FEEDBACK_REPORT_RECIPIENTS", [])
    cc = getattr(settings, "FEEDBACK_REPORT_CC", [])
    if not recipients:
        logger.warning("FEEDBACK_REPORT_RECIPIENTS is empty — skipping email.")
        return "No recipients configured"

    try:
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipients,
            cc=cc,
        )
        email.attach(filename, excel_bytes,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        email.send()
        logger.info("Feedback report sent to %s", ", ".join(recipients))
        return f"Report sent to {', '.join(recipients)}"
    except Exception as exc:
        logger.exception("Failed to send feedback report email")
        raise self.retry(exc=exc)
